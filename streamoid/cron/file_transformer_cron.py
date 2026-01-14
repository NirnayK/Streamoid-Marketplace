import csv
import json
from collections.abc import Iterable
from io import BytesIO, TextIOWrapper
from pathlib import Path
from tempfile import SpooledTemporaryFile

from loguru import logger
from openpyxl import Workbook, load_workbook

from core.file_validation import validate_file_iter
from core.minio import MinioHandler
from mapping.constants import EVALUATION_STATUS_FAILED, EVALUATION_STATUS_SUCCESS, TRANSFORMED_FOLDER
from mapping.models import Mappings
from seller.constants import CSV, MAX_FILE_SIZE, XLSX
from seller.file_parser import FileParser

log = logger.bind(component="file_transformer_cron")


class FileTransformerCron:
    """Stream validation/output generation to avoid loading all rows into memory."""

    @classmethod
    def run(cls):
        pending_mappings = (
            Mappings.objects.filter(evaluation_status="pending")
            .select_related("marketplace_template", "seller_file")
            .iterator()
        )
        for mapping in pending_mappings:
            cls._process_mapping(mapping)

    @classmethod
    def _process_mapping(cls, mapping: Mappings) -> None:
        log_context = log.bind(mapping_id=mapping.id)
        try:
            seller_file = mapping.seller_file
            marketplace_template = mapping.marketplace_template
            if not seller_file or not marketplace_template:
                raise ValueError("Mapping missing seller file or marketplace template.")

            file_bytes = cls._read_file_bytes(seller_file)
            if file_bytes is None:
                raise ValueError("Failed to read seller file from storage.")

            template_keys = list((marketplace_template.template or {}).keys())
            headers, rows = cls._parse_rows(file_bytes, seller_file.file_type)
            # Stream validation/output to keep memory bounded for large files.
            validated_rows = validate_file_iter(marketplace_template, mapping.mappings, headers, rows)
            output_file = cls._build_transformed_file(template_keys, validated_rows, seller_file.file_type)

            bucket_name = seller_file.seller.bucket_name
            object_name = str(Path(TRANSFORMED_FOLDER).joinpath(seller_file.name))
            transformed_path = str(Path(bucket_name).joinpath(object_name))

            try:
                is_stored = MinioHandler().store_file(bucket_name, object_name, output_file)
                if not is_stored:
                    raise ValueError("Failed to store transformed file.")
            finally:
                output_file.close()

            mapping.evaluation_status = EVALUATION_STATUS_SUCCESS
            mapping.transformed_file_path = transformed_path
        except Exception as exc:
            mapping.evaluation_status = EVALUATION_STATUS_FAILED
            mapping.transformed_file_path = None
            log_context.exception("File transformation failed: {}", exc)
        finally:
            mapping.save(update_fields=["evaluation_status", "transformed_file_path", "updated_at"])

    @staticmethod
    def _read_file_bytes(seller_file) -> bytes | None:
        file_object = seller_file.file
        if not file_object:
            return None
        try:
            return file_object.read()
        finally:
            try:
                file_object.close()
            finally:
                release = getattr(file_object, "release_conn", None)
                if callable(release):
                    release()

    @classmethod
    def _parse_rows(cls, file_bytes: bytes, file_type: str) -> tuple[list[str], Iterable[list[str]]]:
        if file_type == CSV:
            return cls._parse_csv_rows(file_bytes)
        if file_type == XLSX:
            return cls._parse_excel_rows(file_bytes)
        raise ValueError("Unsupported file type.")

    @staticmethod
    def _parse_csv_rows(file_bytes: bytes) -> tuple[list[str], Iterable[list[str]]]:
        buffer = BytesIO(file_bytes)
        text_stream = TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
        reader = csv.reader(text_stream)

        header_row = []
        for row in reader:
            row_values = FileParser.normalize_row(row)
            if not any(row_values):
                continue
            header_row = row_values
            break

        if not header_row:
            text_stream.detach()
            buffer.close()
            return [], iter(())

        def row_iter():
            try:
                for row in reader:
                    row_values = FileParser.normalize_row(row)
                    if not any(row_values):
                        continue
                    yield row_values
            finally:
                text_stream.detach()
                buffer.close()

        return header_row, row_iter()

    @staticmethod
    def _parse_excel_rows(file_bytes: bytes) -> tuple[list[str], Iterable[list[str]]]:
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        header_row = []
        for row in rows:
            row_values = FileParser.normalize_row(row)
            if not any(row_values):
                continue
            header_row = row_values
            break

        if not header_row:
            workbook.close()
            return [], iter(())

        def row_iter():
            try:
                for row in rows:
                    row_values = FileParser.normalize_row(row)
                    if not any(row_values):
                        continue
                    yield row_values
            finally:
                workbook.close()

        return header_row, row_iter()

    @classmethod
    def _build_transformed_file(
        cls,
        template_keys: list[str],
        validated_rows: Iterable[dict],
        file_type: str,
    ) -> BytesIO:
        if file_type == CSV:
            return cls._build_csv(template_keys, validated_rows)
        if file_type == XLSX:
            return cls._build_excel(template_keys, validated_rows)
        raise ValueError("Unsupported file type.")

    @staticmethod
    def _build_csv(template_keys: list[str], validated_rows: Iterable[dict]) -> BytesIO:
        output = SpooledTemporaryFile(max_size=MAX_FILE_SIZE, mode="w+b")
        text_stream = TextIOWrapper(output, encoding="utf-8", newline="")
        try:
            writer = csv.writer(text_stream)
            writer.writerow(template_keys)
            for row in validated_rows:
                writer.writerow([FileTransformerCron._format_cell(row.get(key)) for key in template_keys])
            text_stream.flush()
            output.seek(0)
            return output
        finally:
            text_stream.detach()

    @staticmethod
    def _build_excel(template_keys: list[str], validated_rows: Iterable[dict]) -> BytesIO:
        workbook = Workbook(write_only=True)
        output = SpooledTemporaryFile(max_size=MAX_FILE_SIZE, mode="w+b")
        try:
            sheet = workbook.active
            sheet.append(template_keys)
            for row in validated_rows:
                sheet.append([FileTransformerCron._format_cell(row.get(key)) for key in template_keys])
            workbook.save(output)
            output.seek(0)
            return output
        finally:
            workbook.close()

    @staticmethod
    def _format_cell(value):
        if value is None:
            return ""
        if isinstance(value, list):
            return ", ".join(str(item) for item in value)
        if isinstance(value, dict):
            return json.dumps(value)
        return value
