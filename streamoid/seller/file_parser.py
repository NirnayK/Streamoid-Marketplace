import csv
from io import BytesIO, TextIOWrapper

from openpyxl import load_workbook


class FileParser:
    @classmethod
    def normalize_row(cls, row):
        return ["" if cell is None else str(cell).strip() for cell in row]

    @classmethod
    def find_first_non_empty_row(cls, rows):
        for row in rows:
            row_values = cls.normalize_row(row)
            if any(row_values):
                return row_values
        return None

    @classmethod
    def parse_csv(cls, payload_file, sample_limit):
        payload_file.seek(0)
        text_stream = TextIOWrapper(payload_file, encoding="utf-8-sig", newline="")
        try:
            reader = csv.reader(text_stream)
            columns = cls.find_first_non_empty_row(reader) or []
            sample_rows = []
            row_count = 0
            for row in reader:
                row_values = cls.normalize_row(row)
                if not any(row_values):
                    continue
                row_count += 1
                if len(sample_rows) < sample_limit:
                    sample_rows.append(row_values)
            return columns, sample_rows, row_count
        finally:
            text_stream.detach()
            payload_file.seek(0)

    @classmethod
    def parse_excel(cls, payload_file, sample_limit):
        payload_file.seek(0)
        file_bytes = payload_file.read()
        payload_file.seek(0)
        workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            columns = cls.find_first_non_empty_row(rows) or []
            sample_rows = []
            row_count = 0
            for row in rows:
                row_values = cls.normalize_row(row)
                if not any(row_values):
                    continue
                row_count += 1
                if len(sample_rows) < sample_limit:
                    sample_rows.append(row_values)
            return columns, sample_rows, row_count
        finally:
            workbook.close()
