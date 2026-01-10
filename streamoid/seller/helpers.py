import csv
from io import BytesIO, TextIOWrapper

from openpyxl import load_workbook


def normalize_row(row):
    return ["" if cell is None else str(cell).strip() for cell in row]


def find_first_non_empty_row(rows):
    for row in rows:
        row_values = normalize_row(row)
        if any(row_values):
            return row_values
    return None


def parse_csv(payload_file, sample_limit):
    payload_file.seek(0)
    text_stream = TextIOWrapper(payload_file, encoding="utf-8-sig", newline="")
    try:
        reader = csv.reader(text_stream)
        columns = find_first_non_empty_row(reader) or []
        sample_rows = []
        row_count = 0
        for row in reader:
            row_values = normalize_row(row)
            if not any(row_values):
                continue
            row_count += 1
            if len(sample_rows) < sample_limit:
                sample_rows.append(row_values)
        return columns, sample_rows, row_count
    finally:
        text_stream.detach()
        payload_file.seek(0)


def parse_excel(payload_file, sample_limit):
    payload_file.seek(0)
    file_bytes = payload_file.read()
    payload_file.seek(0)
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        columns = find_first_non_empty_row(rows) or []
        sample_rows = []
        row_count = 0
        for row in rows:
            row_values = normalize_row(row)
            if not any(row_values):
                continue
            row_count += 1
            if len(sample_rows) < sample_limit:
                sample_rows.append(row_values)
        return columns, sample_rows, row_count
    finally:
        workbook.close()
