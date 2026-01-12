import csv
from io import BytesIO, TextIOWrapper
from pathlib import Path

from openpyxl import load_workbook
from rest_framework.serializers import FileField, Serializer, ValidationError

from core.constants import MAX_NAME_LENGTH
from core.serializers import ModelSerializerBase
from seller.constants import ALLOWED_EXTENSIONS, CSV, MAX_FILE_SIZE, XLSX
from seller.file_parser import FileParser
from seller.models import Seller, SellerFiles


class SellerSerializer(ModelSerializerBase):
    class Meta:
        model = Seller
        fields = ModelSerializerBase.Meta.fields + ("name",)


class SellerFilesSerializer(ModelSerializerBase):
    seller = SellerSerializer()

    class Meta:
        model = SellerFiles
        fields = ModelSerializerBase.Meta.fields + ("name", "seller", "headers", "rows_count", "sample_rows")


class FileUploadSerialzier(Serializer):
    payload_file = FileField(max_length=MAX_NAME_LENGTH)

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._file_type = None

    def validate_payload_file(self, file):
        self._validate_file_name(file)
        self._validate_file_size(file)
        self._validate_file_type(file)
        self._validate_file_contents(file)
        return file

    def _validate_file_name(self, file):
        name = file.name
        if not name or name.strip() in {".", ".."}:
            raise ValidationError("File name is invalid.")
        if Path(name).name != name:
            raise ValidationError("File name must not include path separators.")

    def _validate_file_type(self, file):
        allowed_extensions = ALLOWED_EXTENSIONS
        extension = Path(file.name).suffix.lower()
        if extension not in allowed_extensions:
            raise ValidationError("Unsupported file type. Only CSV or Excel files are allowed.")

    def _validate_file_size(self, file):
        max_size_bytes = MAX_FILE_SIZE
        if file.size > max_size_bytes:
            raise ValidationError("File size exceeds 10 MB limit.")

    def _validate_file_contents(self, file):
        extension = Path(file.name).suffix.lower()
        if extension == CSV:
            self._file_type = CSV
            self._validate_csv_contents(file)
        else:
            self._file_type = XLSX
            self._validate_excel_contents(file)

    def _validate_csv_contents(self, file):
        sample_bytes = file.read(8192)
        file.seek(0)
        if not sample_bytes:
            raise ValidationError("CSV file is empty.")
        try:
            sample_text = sample_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValidationError("CSV file must be UTF-8 encoded.") from exc
        try:
            csv.Sniffer().sniff(sample_text)
        except csv.Error as exc:
            raise ValidationError("CSV file does not appear to be valid CSV.") from exc
        file.seek(0)
        text_stream = TextIOWrapper(file, encoding="utf-8-sig", newline="")
        try:
            reader = csv.reader(text_stream)
            header_row = FileParser.find_first_non_empty_row(reader)
            if not header_row:
                raise ValidationError("File does not contain any data rows.")
            self._validate_header_row(header_row, "CSV")
        except UnicodeDecodeError as exc:
            raise ValidationError("CSV file must be UTF-8 encoded.") from exc
        finally:
            text_stream.detach()
            file.seek(0)

    def _validate_excel_contents(self, file):
        file_bytes = file.read()
        file.seek(0)
        if not file_bytes:
            raise ValidationError("Excel file is empty.")
        try:
            workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = workbook.active
            header_row = FileParser.find_first_non_empty_row(sheet.iter_rows(values_only=True))
            if not header_row:
                raise ValidationError("File does not contain any data rows.")
            self._validate_header_row(header_row, "Excel")
            workbook.close()
        except Exception as exc:
            raise ValidationError("Excel file could not be read.") from exc

    def _validate_header_row(self, cells, file_type):
        if not cells or any(cell == "" for cell in cells):
            raise ValidationError(f"{file_type} header row must not contain empty columns.")
